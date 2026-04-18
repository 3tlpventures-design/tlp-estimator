"""Build pricing.json from TLP source spreadsheets.

Reads:  data/Allowance_Pricing_Levels.xlsx  (15 category tabs)
        data/BLANK_Allowances_3_5_2024.xlsm (Stairs tab)
Writes: pricing.json (at repo root — served by GitHub Pages)

Usage:  python scripts/build_pricing.py

All tabs share one schema:
    category_id | t1 | t2 | t3 | t4 | t5 | tlpFixed | notes

`tlpFixed` is optional (blank = not fixed). `notes` is freeform and ignored.
Blank `category_id` rows are skipped. Duplicate `category_id` — within a tab,
across tabs, or across workbooks — is a build error.
"""
import json
import sys
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))

from openpyxl import load_workbook

from scripts.config import ALLOWANCE_XLSX, BLANK_XLSM, OUTPUT_PATH, TIERS


def parse_row(row):
    """Return (category_id, entry) or None for a blank row. Raises on bad types."""
    if len(row) < 6:
        raise ValueError(f"row too short: got {len(row)} columns, need >= 6")
    cat_id = row[0]
    if cat_id is None or (isinstance(cat_id, str) and not cat_id.strip()):
        return None
    if not isinstance(cat_id, str):
        raise ValueError(
            f"category_id must be a string; got {type(cat_id).__name__}"
        )
    cat_id = cat_id.strip()
    entry = {tier: row[i + 1] for i, tier in enumerate(TIERS)}
    tlp_fixed = row[6] if len(row) > 6 else None
    if tlp_fixed is True:
        entry["tlpFixed"] = True
    return cat_id, entry


def load_tab(ws):
    """Parse one tab into {category_id: entry}. Errors on duplicate keys."""
    result = {}
    for excel_row, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        parsed = parse_row(row)
        if parsed is None:
            continue
        cat_id, entry = parsed
        if cat_id in result:
            raise ValueError(
                f"duplicate category_id in tab {ws.title!r} "
                f"at row {excel_row}: {cat_id}"
            )
        result[cat_id] = entry
    return result


def load_workbook_pricing(path):
    wb = load_workbook(path, data_only=True, keep_vba=False)
    merged = {}
    for tab_name in wb.sheetnames:
        for cat_id, entry in load_tab(wb[tab_name]).items():
            if cat_id in merged:
                raise ValueError(
                    f"duplicate category_id across tabs in {path.name}: {cat_id}"
                )
            merged[cat_id] = entry
    return merged


def merge_sections(*sections):
    merged = {}
    for section in sections:
        for cat_id, entry in section.items():
            if cat_id in merged:
                raise ValueError(f"duplicate category_id across workbooks: {cat_id}")
            merged[cat_id] = entry
    return merged


def validate_schema(data):
    for cat_id, entry in data.items():
        missing = [t for t in TIERS if t not in entry]
        if missing:
            raise ValueError(f"{cat_id} missing tier keys: {missing}")


def main() -> int:
    allowance = load_workbook_pricing(ALLOWANCE_XLSX)
    blank = load_workbook_pricing(BLANK_XLSM)
    pricing = merge_sections(allowance, blank)
    validate_schema(pricing)

    OUTPUT_PATH.write_text(json.dumps(pricing, indent=2, sort_keys=True) + "\n")

    # Schema-only status output. Never echo raw prices.
    print(f"Wrote {OUTPUT_PATH.name}")
    print(f"Categories: {len(pricing)}")
    coverage = {
        t: sum(1 for e in pricing.values() if e.get(t) is not None)
        for t in TIERS
    }
    print(f"Tier coverage: {coverage}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
